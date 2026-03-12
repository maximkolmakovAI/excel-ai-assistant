import pandas as pd
import openpyxl
import io
import json
from datetime import datetime


SHEET_ALIASES = {
    "Заявки": ["заявки", "заявок", "requests", "applications"],
    "Заказы": ["заказы", "заказов", "orders"],
    "Предприятия": ["предприятия", "предприятий", "enterprises", "companies"],
}

COLUMN_TYPES = {
    "Заявки": {
        "Дата заявки": "date",
        "Номер": "str",
        "Клиент": "str",
        "Распределено": "float",
        "Утверждено": "float",
        "Получено КП": "float",
        "Возвращено ОП": "float",
        "КП клиенту": "float",
        "Заказано": "int",
        "Сумма": "float",
        "Менеджер ОЗ": "str",
    },
    "Заказы": {
        "Дата": "date",
        "Номер": "str",
        "Клиент": "str",
        "Сумма": "float",
        "Валюта": "str",
        "Итоговый доход": "float",
        "Условие": "str",
        "Менеджер ОП": "str",
        "План. дата": "date",
        "Факт. дата": "date",
        "% заказа": "float",
        "% подтв.": "float",
        "% отгр. пост.": "float",
        "% получ.": "float",
        "% отгр. предпр.": "float",
        "Комментарий": "str",
        "Утвержден": "str",
        "Заявка покупателя": "str",
        "Дата создания": "date",
        "Отзыв покупателя": "str",
        "223-ФЗ": "str",
        "Организация": "str",
        "Подразделение": "str",
        "Банк/касса": "str",
        "Спецификация": "str",
    },
    "Предприятия": {
        "Код": "str",
        "Наименование": "str",
        "Полное наименование": "str",
        "ИНН": "str",
        "Регион": "str",
        "ОКВЭД": "str",
        "Дата создания": "date",
    },
}


def _parse_date(val):
    if val is None:
        return pd.NaT
    if isinstance(val, datetime):
        return val
    for fmt in ("%d.%m.%Y", "%Y-%m-%d", "%d/%m/%Y"):
        try:
            return datetime.strptime(str(val).strip(), fmt)
        except Exception:
            pass
    return pd.NaT


def load_excel(file_source) -> dict[str, pd.DataFrame]:
    """Load Excel file. file_source can be a path string or a BytesIO / UploadedFile."""
    if isinstance(file_source, str):
        wb = openpyxl.load_workbook(file_source, data_only=True)
    else:
        content = file_source.read() if hasattr(file_source, "read") else file_source
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)

    sheets: dict[str, pd.DataFrame] = {}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2:
            continue
        # Find header row (first non-empty row)
        header_row_idx = 0
        for i, row in enumerate(rows):
            non_none = [c for c in row if c is not None]
            if len(non_none) >= 2:
                header_row_idx = i
                break
        headers = list(rows[header_row_idx])
        data_rows = rows[header_row_idx + 1:]

        # Strip leading None columns
        start_col = 0
        for idx, h in enumerate(headers):
            if h is not None:
                start_col = idx
                break

        headers = headers[start_col:]
        data_rows = [row[start_col:] for row in data_rows]

        # Fill trailing None headers with auto names
        clean_headers = []
        for idx, h in enumerate(headers):
            if h is None:
                clean_headers.append(f"Колонка_{idx + 1}")
            else:
                clean_headers.append(str(h).strip())

        df = pd.DataFrame(data_rows, columns=clean_headers)
        df = df.dropna(how="all")

        # Apply known type conversions
        type_map = COLUMN_TYPES.get(sheet_name, {})
        for col, ctype in type_map.items():
            if col in df.columns:
                if ctype == "date":
                    df[col] = df[col].apply(_parse_date)
                    df[col] = pd.to_datetime(df[col], errors="coerce")
                elif ctype == "float":
                    df[col] = pd.to_numeric(df[col], errors="coerce")
                elif ctype == "int":
                    df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0).astype(int)
                elif ctype == "str":
                    df[col] = df[col].apply(lambda x: str(x).strip() if x is not None and not pd.isna(x) else None)

        sheets[sheet_name] = df

    return sheets


def get_schema_summary(sheets: dict[str, pd.DataFrame]) -> str:
    """Return a text description of all loaded sheets for the AI prompt."""
    lines = []
    for sheet_name, df in sheets.items():
        lines.append(f"## Лист: {sheet_name} ({len(df)} строк)")
        for col in df.columns:
            dtype = str(df[col].dtype)
            sample_vals = df[col].dropna().head(3).tolist()
            sample_str = ", ".join(str(v) for v in sample_vals)
            lines.append(f"  - {col} ({dtype}): примеры: {sample_str}")
        lines.append("")
    return "\n".join(lines)


def get_sheet_prefix_info(sheets: dict[str, pd.DataFrame]) -> str:
    """Extract prefix info from application numbers to determine branches."""
    info = []
    if "Заявки" in sheets:
        df = sheets["Заявки"]
        if "Номер" in df.columns:
            prefixes = df["Номер"].dropna().str[:3].value_counts().head(10)
            info.append("Префиксы номеров заявок (первые 3 символа = код подразделения/филиала):")
            for prefix, count in prefixes.items():
                info.append(f"  {prefix}: {count} заявок")
    return "\n".join(info)
